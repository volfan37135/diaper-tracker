import os
import io
import logging
from datetime import datetime, date
from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, jsonify, send_file
)
from flask_wtf.csrf import CSRFProtect
from forms import PurchaseForm, BrandForm
import models

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get(
    'SECRET_KEY', 'change-this-to-a-random-secret-key'
)
csrf = CSRFProtect(app)

EXPORT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
os.makedirs(EXPORT_DIR, exist_ok=True)


@app.after_request
def add_no_cache(response):
    if 'text/html' in response.content_type:
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
    return response


@app.before_request
def ensure_db():
    if not hasattr(app, '_db_initialized'):
        models.init_db()
        models.migrate_add_size_column()
        models.migrate_box_openings()
        app._db_initialized = True


# --- Dashboard ---

@app.route('/')
def index():
    stats = models.get_statistics()
    recent = models.get_all_purchases()[:5]
    return render_template('index.html', stats=stats, recent=recent)


# --- Purchases ---

@app.route('/add', methods=['GET', 'POST'])
def add_purchase():
    form = PurchaseForm()
    brands = models.get_brand_names()

    if request.method == 'POST':
        logger.info(f"POST /add - form valid: {form.validate_on_submit()}, errors: {form.errors}")

    if form.validate_on_submit():
        diapers_per_box = form.diapers_per_box.data
        if diapers_per_box == 'custom':
            diapers_per_box = request.form.get('custom_diapers_per_box', '')
            if not diapers_per_box or not diapers_per_box.isdigit() or int(diapers_per_box) < 1:
                flash('Please enter a valid custom quantity.', 'error')
                return render_template('add_purchase.html', form=form, brands=brands)
            diapers_per_box = int(diapers_per_box)
        else:
            diapers_per_box = int(diapers_per_box)

        purchase_id = models.add_purchase(
            date=form.date.data.isoformat(),
            num_boxes=int(form.num_boxes.data),
            diapers_per_box=diapers_per_box,
            brand=form.brand.data.strip(),
            cost=float(form.cost.data),
            size=form.size.data
        )
        if form.date_opened.data:
            openings = models.get_box_openings(purchase_id)
            if openings:
                models.update_box_opening(openings[0]['id'], form.date_opened.data.isoformat())
        logger.info(f"Purchase saved: {form.date.data} {form.brand.data} {form.num_boxes.data} boxes")
        flash('Purchase added successfully!', 'success')
        return redirect(url_for('history'))

    return render_template('add_purchase.html', form=form, brands=brands)


@app.route('/history')
def history():
    purchases = models.get_all_purchases()
    box_openings = models.get_all_box_openings()
    return render_template('history.html', purchases=purchases, box_openings=box_openings, today=date.today().isoformat())


@app.route('/delete_purchase/<int:purchase_id>', methods=['POST'])
def delete_purchase(purchase_id):
    models.delete_purchase(purchase_id)
    flash('Purchase deleted.', 'success')
    return redirect(url_for('history'))


@app.route('/purchase/<int:purchase_id>/open_box', methods=['POST'])
def open_box(purchase_id):
    opening_id = request.form.get('opening_id')
    date_opened = request.form.get('date_opened')
    if opening_id and date_opened:
        models.update_box_opening(int(opening_id), date_opened)
        flash('Box opening date saved.', 'success')
    else:
        flash('Missing opening ID or date.', 'error')
    return redirect(url_for('history'))


# --- Brands ---

@app.route('/brands', methods=['GET', 'POST'])
def manage_brands():
    form = BrandForm()
    if form.validate_on_submit():
        models.add_brand(form.brand_name.data)
        flash('Brand added.', 'success')
        return redirect(url_for('manage_brands'))
    brands = models.get_all_brands()
    return render_template('manage_brands.html', form=form, brands=brands)


@app.route('/brands/edit/<int:brand_id>', methods=['POST'])
def edit_brand(brand_id):
    new_name = request.form.get('brand_name', '').strip()
    if new_name:
        models.update_brand(brand_id, new_name)
        flash('Brand updated.', 'success')
    else:
        flash('Brand name cannot be empty.', 'error')
    return redirect(url_for('manage_brands'))


@app.route('/brands/delete/<int:brand_id>', methods=['POST'])
def delete_brand(brand_id):
    models.delete_brand(brand_id)
    flash('Brand deleted from saved list. Existing purchase records are unchanged.', 'success')
    return redirect(url_for('manage_brands'))


@app.route('/api/brands')
def api_brands():
    return jsonify(models.get_brand_names())


# --- Export PDF ---

@app.route('/export/pdf')
def export_pdf():
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )
    from reportlab.lib.styles import getSampleStyleSheet

    stats = models.get_statistics()
    purchases = models.get_all_purchases()
    box_openings = models.get_all_box_openings()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph('Diaper Purchase Report', styles['Title']))
    elements.append(Paragraph(
        f'Generated: {datetime.now().strftime("%B %d, %Y")}',
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.3 * inch))

    # Stats
    elements.append(Paragraph('Summary Statistics', styles['Heading2']))
    stats_data = [
        ['Total Purchases', str(stats['total_purchases'])],
        ['Total Boxes', str(stats['total_boxes'])],
        ['Total Diapers', f"{stats['total_diapers']:,}"],
        ['Total Cost', f"${stats['total_cost']:,.2f}"],
        ['Avg Cost/Diaper', f"${stats['avg_cost_per_diaper']:.4f}"],
    ]
    t = Table(stats_data, colWidths=[2.5 * inch, 2.5 * inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8f4fd')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 0.3 * inch))

    # Purchases table
    if purchases:
        elements.append(Paragraph('Purchase History', styles['Heading2']))
        header = ['Date', 'Brand', 'Size', 'Boxes', 'Per Box', 'Total', 'Cost', '$/Diaper', 'Opened']
        data = [header]
        for p in purchases:
            total = p['num_boxes'] * p['diapers_per_box']
            cpd = p['cost'] / total if total > 0 else 0
            openings = box_openings.get(p['id'], [])
            opened_count = sum(1 for o in openings if o['date_opened'])
            opened_str = f"{opened_count}/{p['num_boxes']}"
            data.append([
                p['date'],
                p['brand'][:25],
                p['size'] or '',
                str(p['num_boxes']),
                str(p['diapers_per_box']),
                str(total),
                f"${p['cost']:.2f}",
                f"${cpd:.4f}",
                opened_str,
            ])
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(t)

    doc.build(elements)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'diaper_report_{date.today().isoformat()}.pdf'
    )


# --- Export Excel ---

@app.route('/export/excel')
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    stats = models.get_statistics()
    purchases = models.get_all_purchases()
    box_openings = models.get_all_box_openings()

    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    header_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    ws_summary['A1'] = 'Diaper Purchase Report'
    ws_summary['A1'].font = header_font
    ws_summary['A2'] = f'Generated: {datetime.now().strftime("%B %d, %Y")}'

    summary_items = [
        ('Total Purchases', stats['total_purchases']),
        ('Total Boxes', stats['total_boxes']),
        ('Total Diapers', stats['total_diapers']),
        ('Total Cost', f"${stats['total_cost']:,.2f}"),
        ('Avg Cost per Diaper', f"${stats['avg_cost_per_diaper']:.4f}"),
    ]
    for i, (label, value) in enumerate(summary_items, start=4):
        ws_summary[f'A{i}'] = label
        ws_summary[f'A{i}'].font = label_font
        ws_summary[f'B{i}'] = value
    ws_summary.column_dimensions['A'].width = 22
    ws_summary.column_dimensions['B'].width = 18

    # Purchases sheet
    ws_data = wb.create_sheet('Purchases')
    headers = ['Date', 'Brand', 'Size', 'Boxes', 'Diapers/Box', 'Total Diapers', 'Cost', 'Cost/Diaper', 'Boxes Opened']
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_font_white = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, p in enumerate(purchases, 2):
        total = p['num_boxes'] * p['diapers_per_box']
        cpd = p['cost'] / total if total > 0 else 0
        openings = box_openings.get(p['id'], [])
        opened_count = sum(1 for o in openings if o['date_opened'])
        values = [
            p['date'], p['brand'], p['size'] or '', p['num_boxes'], p['diapers_per_box'],
            total, round(p['cost'], 2), round(cpd, 4),
            f"{opened_count}/{p['num_boxes']}"
        ]
        for col, val in enumerate(values, 1):
            cell = ws_data.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border

    for col_letter, width in [('A', 12), ('B', 20), ('C', 10), ('D', 8), ('E', 13), ('F', 14), ('G', 10), ('H', 12), ('I', 14)]:
        ws_data.column_dimensions[col_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'diaper_data_{date.today().isoformat()}.xlsx'
    )


if __name__ == '__main__':
    models.init_db()
    app.run(debug=True, port=5003)
